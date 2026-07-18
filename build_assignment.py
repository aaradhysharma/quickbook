"""
Brew Haven Cafe — Cost Tracker, Profitability & Regression Forecasting
Generates Excel workbook, charts, and regression stats for the assignment.
"""
from pathlib import Path
import json

import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use("Agg")

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.chart import ScatterChart, Reference, Series, BarChart, LineChart
from openpyxl.chart.marker import Marker
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

ROOT = Path(__file__).parent
ASSETS = ROOT / "assets"
SCREENSHOTS = ROOT / "screenshots"
ASSETS.mkdir(exist_ok=True)
SCREENSHOTS.mkdir(exist_ok=True)

VERSION = "0.0.1"

# ---------------------------------------------------------------------------
# Business scenario data: Brew Haven Cafe (hypothetical coffee shop)
# ---------------------------------------------------------------------------

STARTUP_COSTS = [
    ("Espresso machine (2-group)", 8500, "Fixed", "Equipment"),
    ("Commercial grinders (x2)", 1600, "Fixed", "Equipment"),
    ("Furniture & seating", 5200, "Fixed", "Furnishings"),
    ("Lease deposit (2 months)", 7000, "Fixed", "Facilities"),
    ("Build-out / renovations", 9500, "Fixed", "Facilities"),
    ("Business licenses & permits", 850, "Fixed", "Legal"),
    ("Initial inventory (beans, milk, cups)", 2800, "Variable", "Inventory"),
    ("POS system & tablets", 1400, "Fixed", "Technology"),
    ("Exterior signage & branding", 1200, "Fixed", "Marketing"),
    ("Opening marketing campaign", 1500, "Fixed", "Marketing"),
    ("Smallwares (pitchers, scales, etc.)", 650, "Fixed", "Equipment"),
    ("Working capital buffer", 5000, "Fixed", "Cash reserve"),
]

# Monthly ongoing expenses (typical month at ~4,200 cups)
ONGOING = [
    ("Rent", 3500, "Fixed", "Facilities"),
    ("Base salaries (manager + 2 baristas)", 8200, "Fixed", "Labor"),
    ("Hourly labor (volume-driven)", 2100, "Variable", "Labor"),
    ("Coffee beans & tea", 1680, "Variable", "COGS"),
    ("Milk, syrups & dairy alternatives", 920, "Variable", "COGS"),
    ("Cups, lids, sleeves, napkins", 480, "Variable", "COGS"),
    ("Utilities (electricity, water, gas)", 620, "Mixed", "Facilities"),
    ("Insurance", 275, "Fixed", "Insurance"),
    ("Software (POS, payroll, accounting)", 165, "Fixed", "Technology"),
    ("Marketing / social ads", 400, "Fixed", "Marketing"),
    ("Maintenance & supplies", 180, "Fixed", "Operations"),
    ("Loan / equipment payment", 450, "Fixed", "Debt"),
]

# 12 months of historical volume & cost/revenue for regression
# Volume (cups sold), Total Cost, Revenue
MONTHLY_DATA = [
    ("Jan", 3200, 14820, 17600),
    ("Feb", 3450, 15240, 18975),
    ("Mar", 3800, 15890, 20900),
    ("Apr", 4100, 16450, 22550),
    ("May", 4450, 17120, 24475),
    ("Jun", 4800, 17800, 26400),
    ("Jul", 5100, 18350, 28050),
    ("Aug", 4950, 18120, 27225),
    ("Sep", 4300, 16880, 23650),
    ("Oct", 4000, 16290, 22000),
    ("Nov", 3700, 15710, 20350),
    ("Dec", 4600, 17460, 25300),
]

AVG_PRICE = 5.50  # blended avg ticket per cup

# Styles
HEADER_FILL = PatternFill("solid", fgColor="1B4332")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
TITLE_FONT = Font(bold=True, name="Calibri", size=16, color="1B4332")
SUBTITLE_FONT = Font(bold=True, name="Calibri", size=12, color="2D6A4F")
MONEY_FORMAT = '"$"#,##0.00'
INT_FORMAT = "#,##0"
PCT_FORMAT = "0.0%"
THIN = Border(
    left=Side(style="thin", color="D8E2DC"),
    right=Side(style="thin", color="D8E2DC"),
    top=Side(style="thin", color="D8E2DC"),
    bottom=Side(style="thin", color="D8E2DC"),
)
ALT_FILL = PatternFill("solid", fgColor="F0F7F4")
FIXED_FILL = PatternFill("solid", fgColor="E8F5E9")
VAR_FILL = PatternFill("solid", fgColor="FFF3E0")
MIXED_FILL = PatternFill("solid", fgColor="E3F2FD")
PROFIT_FILL = PatternFill("solid", fgColor="C8E6C9")
LOSS_FILL = PatternFill("solid", fgColor="FFCDD2")


def style_header_row(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN


def autosize(ws, min_width=10, max_width=42):
    for col_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


def run_regression(volumes, costs):
    """Simple linear regression: Total Cost = a + b * Volume"""
    x = np.array(volumes, dtype=float)
    y = np.array(costs, dtype=float)
    slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
    y_pred = intercept + slope * x
    residuals = y - y_pred
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((y - np.mean(y)) ** 2)
    r_squared = 1 - ss_res / ss_tot
    # Forecast next 3 months at projected volumes
    forecast_vols = [5200, 5400, 5600]
    forecasts = [(v, intercept + slope * v) for v in forecast_vols]
    return {
        "intercept": float(intercept),
        "slope": float(slope),
        "r_value": float(r_value),
        "r_squared": float(r_squared),
        "p_value": float(p_value),
        "std_err": float(std_err),
        "n": len(x),
        "forecasts": forecasts,
        "y_pred": y_pred.tolist(),
    }


def make_charts(volumes, costs, revenues, reg):
    # 1) Cost vs Volume scatter + regression line
    fig, ax = plt.subplots(figsize=(9, 5.5), facecolor="#FAFBF9")
    ax.set_facecolor("#FAFBF9")
    ax.scatter(volumes, costs, s=80, c="#2D6A4F", zorder=3, label="Actual monthly total cost")
    x_line = np.linspace(min(volumes) - 200, max(volumes) + 400, 100)
    y_line = reg["intercept"] + reg["slope"] * x_line
    ax.plot(x_line, y_line, color="#BC4749", linewidth=2.2, label=f"Regression: TC = {reg['intercept']:.0f} + {reg['slope']:.2f}×Q")
    for vol, cost in reg["forecasts"]:
        ax.scatter([vol], [cost], s=100, c="#E09F3E", marker="D", zorder=4)
    ax.scatter([], [], s=100, c="#E09F3E", marker="D", label="Forecast points")
    ax.set_xlabel("Cups sold (monthly volume)", fontsize=11)
    ax.set_ylabel("Total cost ($)", fontsize=11)
    ax.set_title("Brew Haven Cafe — Cost Regression (Total Cost vs Volume)", fontsize=13, fontweight="bold", color="#1B4332")
    ax.legend(loc="upper left", frameon=False)
    ax.grid(True, alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    path1 = SCREENSHOTS / "01_regression_cost_vs_volume.png"
    fig.savefig(path1, dpi=160, bbox_inches="tight")
    plt.close(fig)

    # 2) Monthly P&L bars
    months = [m[0] for m in MONTHLY_DATA]
    fig, ax = plt.subplots(figsize=(10, 5.2), facecolor="#FAFBF9")
    ax.set_facecolor("#FAFBF9")
    x = np.arange(len(months))
    w = 0.35
    ax.bar(x - w / 2, revenues, w, label="Revenue", color="#2D6A4F")
    ax.bar(x + w / 2, costs, w, label="Total Cost", color="#BC4749")
    profit = np.array(revenues) - np.array(costs)
    ax.plot(x, profit, color="#E09F3E", marker="o", linewidth=2, label="Profit")
    ax.set_xticks(x)
    ax.set_xticklabels(months)
    ax.set_ylabel("Dollars ($)")
    ax.set_title("Brew Haven Cafe — Monthly Revenue, Cost & Profit", fontsize=13, fontweight="bold", color="#1B4332")
    ax.legend(frameon=False)
    ax.grid(True, axis="y", alpha=0.3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()
    path2 = SCREENSHOTS / "02_monthly_profitability.png"
    fig.savefig(path2, dpi=160, bbox_inches="tight")
    plt.close(fig)

    # 3) Fixed vs Variable pie (startup + typical month ongoing)
    startup_fixed = sum(c for _, c, t, _ in STARTUP_COSTS if t == "Fixed")
    startup_var = sum(c for _, c, t, _ in STARTUP_COSTS if t == "Variable")
    ongoing_fixed = sum(c for _, c, t, _ in ONGOING if t == "Fixed")
    ongoing_var = sum(c for _, c, t, _ in ONGOING if t == "Variable")
    ongoing_mixed = sum(c for _, c, t, _ in ONGOING if t == "Mixed")

    fig, axes = plt.subplots(1, 2, figsize=(10, 4.8), facecolor="#FAFBF9")
    axes[0].pie(
        [startup_fixed, startup_var],
        labels=["Fixed", "Variable"],
        autopct="%1.1f%%",
        colors=["#2D6A4F", "#E09F3E"],
        startangle=90,
        textprops={"fontsize": 10},
    )
    axes[0].set_title("Startup Costs Mix", fontweight="bold", color="#1B4332")
    axes[1].pie(
        [ongoing_fixed, ongoing_var, ongoing_mixed],
        labels=["Fixed", "Variable", "Mixed"],
        autopct="%1.1f%%",
        colors=["#2D6A4F", "#E09F3E", "#40916C"],
        startangle=90,
        textprops={"fontsize": 10},
    )
    axes[1].set_title("Ongoing Monthly Cost Mix", fontweight="bold", color="#1B4332")
    fig.suptitle("Cost Structure — Fixed vs Variable", fontsize=13, fontweight="bold", color="#1B4332")
    fig.tight_layout()
    path3 = SCREENSHOTS / "03_cost_structure.png"
    fig.savefig(path3, dpi=160, bbox_inches="tight")
    plt.close(fig)

    # 4) Dashboard-style summary "screenshot"
    fig = plt.figure(figsize=(11, 7), facecolor="#1B4332")
    fig.text(0.05, 0.92, "Brew Haven Cafe — Expense Tracker (Excel/Sheets Alternative)", fontsize=16, color="white", fontweight="bold")
    fig.text(0.05, 0.87, "Hands-on Practice Dashboard  |  Tool: Excel via openpyxl  |  Scenario: Neighborhood coffee shop", fontsize=10, color="#95D5B2")

    # Summary boxes
    boxes = [
        (0.05, 0.55, "Startup Capital", f"${sum(c for _, c, _, _ in STARTUP_COSTS):,.0f}", "#2D6A4F"),
        (0.35, 0.55, "Monthly Fixed Costs", f"${ongoing_fixed:,.0f}", "#40916C"),
        (0.65, 0.55, "Monthly Variable (typ.)", f"${ongoing_var:,.0f}", "#E09F3E"),
        (0.05, 0.22, "Avg Monthly Profit", f"${np.mean(profit):,.0f}", "#52B788"),
        (0.35, 0.22, "Variable Cost / Cup", f"${reg['slope']:.2f}", "#BC4749"),
        (0.65, 0.22, "R² (Cost Model)", f"{reg['r_squared']:.3f}", "#95D5B2"),
    ]
    for x, y, label, value, color in boxes:
        axb = fig.add_axes([x, y, 0.28, 0.25])
        axb.set_facecolor(color)
        axb.set_xticks([])
        axb.set_yticks([])
        for spine in axb.spines.values():
            spine.set_visible(False)
        axb.text(0.5, 0.62, value, ha="center", va="center", fontsize=22, fontweight="bold", color="white")
        axb.text(0.5, 0.28, label, ha="center", va="center", fontsize=11, color="white")

    fig.text(0.05, 0.08, f"Regression: Total Cost = ${reg['intercept']:,.0f} + ${reg['slope']:.2f} × Cups   |   p-value = {reg['p_value']:.2e}   |   Version {VERSION}", fontsize=9, color="#D8F3DC")
    path4 = SCREENSHOTS / "04_hands_on_dashboard.png"
    fig.savefig(path4, dpi=160, bbox_inches="tight")
    plt.close(fig)

    # 5) Regression output table image
    fig, ax = plt.subplots(figsize=(8.5, 5), facecolor="#FAFBF9")
    ax.axis("off")
    ax.set_title("Regression Output — Excel / Sheets Style", fontsize=14, fontweight="bold", color="#1B4332", pad=20)
    table_data = [
        ["Statistic", "Value", "Interpretation"],
        ["Intercept (a)", f"${reg['intercept']:,.2f}", "Estimated fixed cost when volume = 0"],
        ["Slope (b)", f"${reg['slope']:.4f}", "Variable cost per additional cup sold"],
        ["R-squared", f"{reg['r_squared']:.4f}", "Share of cost variation explained by volume"],
        ["Correlation (r)", f"{reg['r_value']:.4f}", "Strength of linear volume–cost link"],
        ["p-value", f"{reg['p_value']:.4e}", "Significance of the slope coefficient"],
        ["Std. Error", f"{reg['std_err']:.4f}", "Uncertainty around the slope estimate"],
        ["n (months)", str(reg["n"]), "Sample size used in the model"],
        ["Forecast Q=5,200", f"${reg['forecasts'][0][1]:,.0f}", "Projected total cost at 5,200 cups"],
        ["Forecast Q=5,400", f"${reg['forecasts'][1][1]:,.0f}", "Projected total cost at 5,400 cups"],
        ["Forecast Q=5,600", f"${reg['forecasts'][2][1]:,.0f}", "Projected total cost at 5,600 cups"],
    ]
    table = ax.table(cellText=table_data, loc="center", cellLoc="left", colWidths=[0.22, 0.2, 0.5])
    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1.15, 1.55)
    for j in range(3):
        table[0, j].set_facecolor("#1B4332")
        table[0, j].set_text_props(color="white", fontweight="bold")
    for i in range(1, len(table_data)):
        for j in range(3):
            table[i, j].set_facecolor("#F0F7F4" if i % 2 else "white")
    fig.tight_layout()
    path5 = SCREENSHOTS / "05_regression_output.png"
    fig.savefig(path5, dpi=160, bbox_inches="tight")
    plt.close(fig)

    return {
        "regression_chart": str(path1.name),
        "profitability_chart": str(path2.name),
        "cost_structure_chart": str(path3.name),
        "dashboard": str(path4.name),
        "regression_output": str(path5.name),
    }


def build_workbook(reg):
    wb = Workbook()

    # --- Cover ---
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "Brew Haven Cafe — Cost Tracker & Profitability Analysis"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Alternative to QuickBooks: Excel workbook (Google Sheets–compatible)"
    ws["A2"].font = SUBTITLE_FONT
    ws["A3"] = f"Version {VERSION}  |  Hypothetical neighborhood coffee shop"
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")

    overview = [
        ("Business", "Brew Haven Cafe — specialty coffee shop in a mid-size city"),
        ("Tool used", "Microsoft Excel (openpyxl) — same structure works in Google Sheets"),
        ("Learning goal", "Categorize fixed vs variable costs; project profit; forecast with regression"),
        ("Avg selling price / cup", f"${AVG_PRICE:.2f} blended ticket"),
        ("Startup capital required", sum(c for _, c, _, _ in STARTUP_COSTS)),
        ("Typical monthly fixed costs", sum(c for _, c, t, _ in ONGOING if t == "Fixed")),
        ("Typical monthly variable costs", sum(c for _, c, t, _ in ONGOING if t == "Variable")),
        ("Regression model", f"Total Cost = {reg['intercept']:.2f} + {reg['slope']:.4f} × Cups"),
        ("R-squared", reg["r_squared"]),
    ]
    ws["A5"] = "Field"
    ws["B5"] = "Detail"
    style_header_row(ws, 5, 1, 2)
    for i, (k, v) in enumerate(overview, start=6):
        ws.cell(row=i, column=1, value=k).border = THIN
        cell = ws.cell(row=i, column=2, value=v)
        cell.border = THIN
        if isinstance(v, float) and "R-squared" in k:
            cell.number_format = "0.0000"
        elif isinstance(v, (int, float)) and "cost" in k.lower() or "capital" in k.lower():
            if isinstance(v, (int, float)):
                cell.number_format = MONEY_FORMAT
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = ALT_FILL
            ws.cell(row=i, column=2).fill = ALT_FILL

    ws["A17"] = "How to use (mirrors Jeremy’s Tutorials / Davie Mach spreadsheet bookkeeping)"
    ws["A17"].font = SUBTITLE_FONT
    steps = [
        "1. Enter every cost on Startup Costs and Ongoing Expenses sheets; tag Fixed / Variable / Mixed.",
        "2. Review Monthly Ledger for 12 months of volume, cost, and revenue.",
        "3. Check Profitability for contribution margin, break-even, and projected profit.",
        "4. Open Regression sheet for slope, intercept, R², and 3-month cost forecasts.",
        "5. Use charts / screenshot images for the written report appendix.",
    ]
    for i, s in enumerate(steps, start=18):
        ws.cell(row=i, column=1, value=s)
    autosize(ws)

    # --- Startup Costs ---
    ws = wb.create_sheet("Startup Costs")
    ws["A1"] = "Startup Costs — Brew Haven Cafe"
    ws["A1"].font = TITLE_FONT
    headers = ["#", "Item", "Amount ($)", "Cost Type", "Category"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 1, 5)
    total_startup = 0
    for i, (item, amt, ctype, cat) in enumerate(STARTUP_COSTS, start=1):
        r = 3 + i
        ws.cell(row=r, column=1, value=i).border = THIN
        ws.cell(row=r, column=2, value=item).border = THIN
        cell = ws.cell(row=r, column=3, value=amt)
        cell.number_format = MONEY_FORMAT
        cell.border = THIN
        tcell = ws.cell(row=r, column=4, value=ctype)
        tcell.border = THIN
        tcell.fill = FIXED_FILL if ctype == "Fixed" else VAR_FILL
        ws.cell(row=r, column=5, value=cat).border = THIN
        total_startup += amt
    total_row = 4 + len(STARTUP_COSTS)
    ws.cell(row=total_row, column=2, value="TOTAL STARTUP CAPITAL").font = Font(bold=True)
    t = ws.cell(row=total_row, column=3, value=total_startup)
    t.number_format = MONEY_FORMAT
    t.font = Font(bold=True)
    t.fill = PROFIT_FILL
    ws.cell(row=total_row + 2, column=1, value="Note: Fixed startup costs do not change with first-month sales volume. Initial inventory is treated as variable because usage scales with cups sold.")
    autosize(ws)

    # --- Ongoing Expenses ---
    ws = wb.create_sheet("Ongoing Expenses")
    ws["A1"] = "Ongoing Monthly Expenses — Categorized Fixed / Variable / Mixed"
    ws["A1"].font = TITLE_FONT
    headers = ["#", "Expense", "Amount ($)", "Cost Type", "Category", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 1, 6)
    notes = {
        "Rent": "Lease — constant each month",
        "Base salaries (manager + 2 baristas)": "Guaranteed hours regardless of traffic",
        "Hourly labor (volume-driven)": "Extra shifts when cup volume rises",
        "Coffee beans & tea": "~$0.40/cup COGS",
        "Milk, syrups & dairy alternatives": "~$0.22/cup COGS",
        "Cups, lids, sleeves, napkins": "~$0.11/cup packaging",
        "Utilities (electricity, water, gas)": "Base + usage; partly volume-sensitive",
        "Insurance": "Annual policy / 12",
        "Software (POS, payroll, accounting)": "Subscriptions (QuickBooks-style alternative)",
        "Marketing / social ads": "Planned monthly spend",
        "Maintenance & supplies": "Cleaning & small repairs",
        "Loan / equipment payment": "Financed espresso equipment",
    }
    total_ong = 0
    fixed_sum = var_sum = mixed_sum = 0
    for i, (item, amt, ctype, cat) in enumerate(ONGOING, start=1):
        r = 3 + i
        ws.cell(row=r, column=1, value=i).border = THIN
        ws.cell(row=r, column=2, value=item).border = THIN
        cell = ws.cell(row=r, column=3, value=amt)
        cell.number_format = MONEY_FORMAT
        cell.border = THIN
        tcell = ws.cell(row=r, column=4, value=ctype)
        tcell.border = THIN
        if ctype == "Fixed":
            tcell.fill = FIXED_FILL
            fixed_sum += amt
        elif ctype == "Variable":
            tcell.fill = VAR_FILL
            var_sum += amt
        else:
            tcell.fill = MIXED_FILL
            mixed_sum += amt
        ws.cell(row=r, column=5, value=cat).border = THIN
        ws.cell(row=r, column=6, value=notes.get(item, "")).border = THIN
        total_ong += amt
    r = 4 + len(ONGOING)
    for label, val, fill in [
        ("Total Fixed", fixed_sum, FIXED_FILL),
        ("Total Variable", var_sum, VAR_FILL),
        ("Total Mixed", mixed_sum, MIXED_FILL),
        ("TOTAL MONTHLY", total_ong, PROFIT_FILL),
    ]:
        ws.cell(row=r, column=2, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=3, value=val)
        c.number_format = MONEY_FORMAT
        c.font = Font(bold=True)
        c.fill = fill
        r += 1
    autosize(ws)

    # --- Monthly Ledger ---
    ws = wb.create_sheet("Monthly Ledger")
    ws["A1"] = "12-Month Operating Ledger (Sample Actuals)"
    ws["A1"].font = TITLE_FONT
    headers = ["Month", "Cups Sold (Q)", "Total Cost ($)", "Revenue ($)", "Profit ($)", "Cost/Cup", "Profit Margin"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 1, 7)
    for i, (month, vol, cost, rev) in enumerate(MONTHLY_DATA, start=1):
        r = 3 + i
        profit = rev - cost
        ws.cell(row=r, column=1, value=month).border = THIN
        ws.cell(row=r, column=2, value=vol).border = THIN
        ws.cell(row=r, column=2).number_format = INT_FORMAT
        for col, val in [(3, cost), (4, rev), (5, profit)]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.number_format = MONEY_FORMAT
            cell.border = THIN
            if col == 5:
                cell.fill = PROFIT_FILL if val >= 0 else LOSS_FILL
        cpc = ws.cell(row=r, column=6, value=cost / vol)
        cpc.number_format = MONEY_FORMAT
        cpc.border = THIN
        pm = ws.cell(row=r, column=7, value=profit / rev)
        pm.number_format = PCT_FORMAT
        pm.border = THIN
    # Excel charts
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Monthly Revenue vs Total Cost"
    chart.y_axis.title = "Dollars"
    data = Reference(ws, min_col=3, min_row=3, max_col=4, max_row=15)
    cats = Reference(ws, min_col=1, min_row=4, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 15
    chart.height = 8
    ws.add_chart(chart, "A18")
    autosize(ws)

    # --- Profitability ---
    ws = wb.create_sheet("Profitability")
    ws["A1"] = "Profitability Projection"
    ws["A1"].font = TITLE_FONT
    vols = [m[1] for m in MONTHLY_DATA]
    costs = [m[2] for m in MONTHLY_DATA]
    revs = [m[3] for m in MONTHLY_DATA]
    avg_vol = np.mean(vols)
    avg_cost = np.mean(costs)
    avg_rev = np.mean(revs)
    avg_profit = avg_rev - avg_cost
    # From regression: VC = slope, FC = intercept
    vc = reg["slope"]
    fc = reg["intercept"]
    cm_per_cup = AVG_PRICE - vc
    break_even_cups = fc / cm_per_cup if cm_per_cup > 0 else None

    rows = [
        ("Average monthly cups", avg_vol, INT_FORMAT),
        ("Average monthly revenue", avg_rev, MONEY_FORMAT),
        ("Average monthly total cost", avg_cost, MONEY_FORMAT),
        ("Average monthly profit", avg_profit, MONEY_FORMAT),
        ("Average profit margin", avg_profit / avg_rev, PCT_FORMAT),
        ("Estimated fixed cost (regression intercept)", fc, MONEY_FORMAT),
        ("Estimated variable cost per cup (slope)", vc, MONEY_FORMAT),
        ("Average selling price per cup", AVG_PRICE, MONEY_FORMAT),
        ("Contribution margin per cup", cm_per_cup, MONEY_FORMAT),
        ("Break-even volume (cups / month)", break_even_cups, INT_FORMAT),
    ]
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    style_header_row(ws, 3, 1, 2)
    for i, (label, val, fmt) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).border = THIN
        cell = ws.cell(row=i, column=2, value=val)
        cell.number_format = fmt
        cell.border = THIN
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = ALT_FILL
            cell.fill = ALT_FILL

    ws["A16"] = "Scenario projections (using regression cost model + $5.50 avg price)"
    ws["A16"].font = SUBTITLE_FONT
    headers = ["Scenario", "Cups / Month", "Projected Cost", "Projected Revenue", "Projected Profit"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=17, column=c, value=h)
    style_header_row(ws, 17, 1, 5)
    scenarios = [("Conservative", 3500), ("Base", 4200), ("Growth", 5200), ("Peak summer", 5600)]
    for i, (name, q) in enumerate(scenarios, start=18):
        cost = fc + vc * q
        rev = AVG_PRICE * q
        profit = rev - cost
        ws.cell(row=i, column=1, value=name).border = THIN
        ws.cell(row=i, column=2, value=q).border = THIN
        for col, val in [(3, cost), (4, rev), (5, profit)]:
            cell = ws.cell(row=i, column=col, value=val)
            cell.number_format = MONEY_FORMAT
            cell.border = THIN
            if col == 5:
                cell.fill = PROFIT_FILL if profit >= 0 else LOSS_FILL
    autosize(ws)

    # --- Regression ---
    ws = wb.create_sheet("Regression")
    ws["A1"] = "Simple Linear Regression — Forecasting Total Cost from Volume"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Model: Total Cost (Y) = Intercept (a) + Slope (b) × Cups Sold (X)"
    ws["A2"].font = SUBTITLE_FONT

    ws["A4"] = "Regression Output (equivalent to Excel Data Analysis > Regression)"
    ws["A4"].font = Font(bold=True)
    headers = ["Statistic", "Value", "Interpretation"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=5, column=c, value=h)
    style_header_row(ws, 5, 1, 3)
    output_rows = [
        ("Intercept (a) — Fixed cost estimate", reg["intercept"], f"When cups = 0, expected monthly cost ≈ ${reg['intercept']:,.0f}"),
        ("Slope (b) — Variable cost / cup", reg["slope"], f"Each extra cup adds ≈ ${reg['slope']:.2f} to total cost"),
        ("R-squared", reg["r_squared"], f"{reg['r_squared']*100:.1f}% of cost variation explained by volume"),
        ("Correlation (r)", reg["r_value"], "Strong positive linear relationship"),
        ("p-value (slope)", reg["p_value"], "Slope is statistically significant (p < 0.05)"),
        ("Standard error (slope)", reg["std_err"], "Precision of the variable-cost estimate"),
        ("Observations (n)", reg["n"], "Months of operating data"),
    ]
    for i, (stat, val, interp) in enumerate(output_rows, start=6):
        ws.cell(row=i, column=1, value=stat).border = THIN
        cell = ws.cell(row=i, column=2, value=val)
        cell.border = THIN
        if isinstance(val, float) and val < 1 and "p-value" not in stat.lower() and "R-squared" in stat or "Correlation" in stat or "R-squared" in stat:
            cell.number_format = "0.0000"
        elif isinstance(val, float) and ("Intercept" in stat or "Slope" in stat or "Standard" in stat):
            cell.number_format = MONEY_FORMAT if "Intercept" in stat or "Slope" in stat else "0.0000"
        elif isinstance(val, float) and "p-value" in stat.lower():
            cell.number_format = "0.00E+00"
        elif isinstance(val, (int, float)) and "Observations" in stat:
            cell.number_format = INT_FORMAT
        else:
            if isinstance(val, float):
                cell.number_format = "0.0000"
        ws.cell(row=i, column=3, value=interp).border = THIN

    ws["A15"] = "Fitted values & residuals"
    ws["A15"].font = Font(bold=True)
    headers = ["Month", "Cups (X)", "Actual Cost (Y)", "Predicted Cost", "Residual"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=16, column=c, value=h)
    style_header_row(ws, 16, 1, 5)
    for i, ((month, vol, cost, _), yhat) in enumerate(zip(MONTHLY_DATA, reg["y_pred"]), start=17):
        ws.cell(row=i, column=1, value=month).border = THIN
        ws.cell(row=i, column=2, value=vol).border = THIN
        for col, val in [(3, cost), (4, yhat), (5, cost - yhat)]:
            cell = ws.cell(row=i, column=col, value=val)
            cell.number_format = MONEY_FORMAT
            cell.border = THIN

    ws["A31"] = "3-Month Cost Forecast"
    ws["A31"].font = Font(bold=True)
    headers = ["Period", "Projected Cups", "Forecast Total Cost", "Implied Revenue @ $5.50", "Implied Profit"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=32, column=c, value=h)
    style_header_row(ws, 32, 1, 5)
    for i, ((vol, cost), label) in enumerate(zip(reg["forecasts"], ["Month +1", "Month +2", "Month +3"]), start=33):
        rev = AVG_PRICE * vol
        profit = rev - cost
        ws.cell(row=i, column=1, value=label).border = THIN
        ws.cell(row=i, column=2, value=vol).border = THIN
        for col, val in [(3, cost), (4, rev), (5, profit)]:
            cell = ws.cell(row=i, column=col, value=val)
            cell.number_format = MONEY_FORMAT
            cell.border = THIN
            if col == 5:
                cell.fill = PROFIT_FILL

    # Scatter chart in Excel
    # Put X,Y for chart in columns G:H
    ws["G16"] = "Cups"
    ws["H16"] = "Cost"
    for i, (_, vol, cost, _) in enumerate(MONTHLY_DATA, start=17):
        ws.cell(row=i, column=7, value=vol)
        ws.cell(row=i, column=8, value=cost)
    scatter = ScatterChart()
    scatter.title = "Total Cost vs Cups Sold"
    scatter.x_axis.title = "Cups"
    scatter.y_axis.title = "Total Cost ($)"
    xvalues = Reference(ws, min_col=7, min_row=17, max_row=28)
    yvalues = Reference(ws, min_col=8, min_row=16, max_row=28)
    series = Series(yvalues, xvalues, title="Actual")
    scatter.series.append(series)
    scatter.width = 14
    scatter.height = 9
    ws.add_chart(scatter, "A38")

    # Embed regression chart image if available
    img_path = SCREENSHOTS / "01_regression_cost_vs_volume.png"
    if img_path.exists():
        img = XLImage(str(img_path))
        img.width = 540
        img.height = 330
        ws.add_image(img, "G38")

    autosize(ws)
    ws.column_dimensions["C"].width = 55

    # --- Cost Type Guide ---
    ws = wb.create_sheet("Cost Type Guide")
    ws["A1"] = "Fixed vs Variable Costs — Quick Reference"
    ws["A1"].font = TITLE_FONT
    guide = [
        ("Fixed costs", "Stay the same within a relevant range of activity (rent, salaried managers, insurance, software subscriptions)."),
        ("Variable costs", "Change in direct proportion to volume (beans, milk, cups, hourly labor tied to shifts)."),
        ("Mixed (semi-variable)", "Contain a fixed base plus a usage component (utilities)."),
        ("Why it matters for pricing", "Price must cover variable cost per unit AND contribute toward fixed costs and target profit."),
        ("Contribution margin", "Selling price − variable cost per unit. Used for break-even and pricing floors."),
        ("Break-even cups", "Fixed costs ÷ contribution margin per cup."),
    ]
    ws["A3"] = "Concept"
    ws["B3"] = "Explanation"
    style_header_row(ws, 3, 1, 2)
    for i, (a, b) in enumerate(guide, start=4):
        ws.cell(row=i, column=1, value=a).border = THIN
        ws.cell(row=i, column=1).fill = FIXED_FILL
        ws.cell(row=i, column=2, value=b).border = THIN
    autosize(ws)

    out = ROOT / "Brew_Haven_Cafe_Cost_Tracker.xlsx"
    wb.save(out)
    return out, total_startup, fixed_sum, var_sum, mixed_sum, break_even_cups, cm_per_cup, avg_profit


def main():
    volumes = [m[1] for m in MONTHLY_DATA]
    costs = [m[2] for m in MONTHLY_DATA]
    revenues = [m[3] for m in MONTHLY_DATA]
    reg = run_regression(volumes, costs)
    chart_files = make_charts(volumes, costs, revenues, reg)
    xlsx_path, total_startup, fixed_sum, var_sum, mixed_sum, be, cm, avg_profit = build_workbook(reg)

    summary = {
        "version": VERSION,
        "business": "Brew Haven Cafe",
        "tool": "Excel (Google Sheets compatible) — QuickBooks alternative",
        "startup_total": total_startup,
        "monthly_fixed": fixed_sum,
        "monthly_variable": var_sum,
        "monthly_mixed": mixed_sum,
        "regression": {
            "intercept": reg["intercept"],
            "slope": reg["slope"],
            "r_squared": reg["r_squared"],
            "r_value": reg["r_value"],
            "p_value": reg["p_value"],
            "std_err": reg["std_err"],
            "n": reg["n"],
            "forecasts": [{"volume": v, "cost": c} for v, c in reg["forecasts"]],
            "equation": f"Total Cost = {reg['intercept']:.2f} + {reg['slope']:.4f} × Cups",
        },
        "break_even_cups": be,
        "contribution_margin": cm,
        "avg_monthly_profit": float(avg_profit),
        "avg_price": AVG_PRICE,
        "charts": chart_files,
        "workbook": xlsx_path.name,
    }
    (ASSETS / "summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    print(f"\nWorkbook: {xlsx_path}")
    print(f"Charts in: {SCREENSHOTS}")


if __name__ == "__main__":
    main()
